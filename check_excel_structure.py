import openpyxl
import os

def check_excel_structure():
    """엑셀 파일의 실제 구조 확인"""
    excel_file = "와석초 카카오톡 챗봇 개발을 위한 질문과 답변(0702).xlsx"
    
    if not os.path.exists(excel_file):
        print(f"엑셀 파일을 찾을 수 없습니다: {excel_file}")
        return
    
    wb = openpyxl.load_workbook(excel_file)
    
    for sheet_name in wb.sheetnames:
        print(f"\n=== {sheet_name} 시트 ===")
        ws = wb[sheet_name]
        
        # 헤더 확인
        print("헤더 (첫 번째 행):")
        for col in range(1, min(6, ws.max_column + 1)):
            cell_value = ws.cell(row=1, column=col).value
            print(f"  열 {col}: {cell_value}")
        
        # 실제 데이터 샘플 (2-5행)
        print("\n데이터 샘플 (2-5행):")
        for row in range(2, min(6, ws.max_row + 1)):
            print(f"\n행 {row}:")
            for col in range(1, min(6, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    print(f"  열 {col}: {cell_value}")
    
    wb.close()

if __name__ == "__main__":
    check_excel_structure() 