import openpyxl
import os

def check_excel_structure():
    """엑셀 파일의 구조를 확인"""
    excel_file = "와석초 카카오톡 챗봇 개발을 위한 질문과 답변(0702).xlsx"
    
    if not os.path.exists(excel_file):
        print(f"엑셀 파일을 찾을 수 없습니다: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        print(f"엑셀 파일: {excel_file}")
        print(f"시트 목록: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n=== {sheet_name} 시트 ===")
            ws = wb[sheet_name]
            
            # 첫 5행 확인
            for row_num in range(1, min(6, ws.max_row + 1)):
                row_data = []
                for col_num in range(1, min(6, ws.max_column + 1)):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data.append(str(cell_value) if cell_value is not None else "None")
                print(f"행 {row_num}: {row_data}")
            
            print(f"총 행 수: {ws.max_row}, 총 열 수: {ws.max_column}")
        
        wb.close()
        
    except Exception as e:
        print(f"엑셀 파일 확인 중 오류: {e}")

if __name__ == "__main__":
    check_excel_structure() 