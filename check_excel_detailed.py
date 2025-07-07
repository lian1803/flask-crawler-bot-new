import openpyxl
import os

def check_excel_detailed():
    """엑셀 파일의 모든 시트를 자세히 확인"""
    excel_file = "와석초 카카오톡 챗봇 개발을 위한 질문과 답변(0702).xlsx"
    
    if not os.path.exists(excel_file):
        print(f"엑셀 파일을 찾을 수 없습니다: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        print(f"엑셀 파일: {excel_file}")
        print(f"시트 목록: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*50}")
            print(f"=== {sheet_name} 시트 ===")
            print(f"{'='*50}")
            ws = wb[sheet_name]
            
            print(f"총 행 수: {ws.max_row}, 총 열 수: {ws.max_column}")
            
            # 모든 행 확인 (빈 행 제외)
            data_rows = 0
            for row_num in range(1, ws.max_row + 1):
                row_data = []
                has_data = False
                
                for col_num in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value is not None and str(cell_value).strip():
                        has_data = True
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                if has_data:
                    data_rows += 1
                    print(f"행 {row_num}: {row_data}")
                    
                    # 최대 20행까지만 출력
                    if data_rows >= 20:
                        print(f"... (총 {ws.max_row}행 중 20행만 표시)")
                        break
        
        wb.close()
        
    except Exception as e:
        print(f"엑셀 파일 확인 중 오류: {e}")

if __name__ == "__main__":
    check_excel_detailed() 