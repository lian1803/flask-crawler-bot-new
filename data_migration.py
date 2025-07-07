import sqlite3
import openpyxl
import os
import shutil
from datetime import datetime

def migrate_data():
    """기존 rol 폴더의 데이터를 new 폴더로 마이그레이션"""
    print("데이터 마이그레이션 시작...")
    
    # 1. 기존 데이터베이스 파일 복사
    source_db = "../rol/school_data.db"
    target_db = "school_data.db"
    
    if os.path.exists(source_db):
        shutil.copy2(source_db, target_db)
        print(f"데이터베이스 파일 복사 완료: {source_db} -> {target_db}")
    else:
        print(f"기존 데이터베이스 파일을 찾을 수 없습니다: {source_db}")
    
    # 2. 엑셀 파일 복사 (새로운 파일이 있으면 복사하지 않음)
    source_excel = "../rol/와석초 카카오톡 챗봇 개발을 위한 질문과 답변 의 사본.xlsx"
    target_excel = "와석초 카카오톡 챗봇 개발을 위한 질문과 답변 의 사본.xlsx"
    
    # 새로운 엑셀 파일이 이미 있으면 복사하지 않음
    if not os.path.exists("와석초 카카오톡 챗봇 개발을 위한 질문과 답변(0702).xlsx"):
        if os.path.exists(source_excel):
            shutil.copy2(source_excel, target_excel)
            print(f"엑셀 파일 복사 완료: {source_excel} -> {target_excel}")
        else:
            print(f"기존 엑셀 파일을 찾을 수 없습니다: {source_excel}")
    else:
        print("새로운 엑셀 파일이 이미 존재합니다.")
    
    if os.path.exists(source_excel):
        shutil.copy2(source_excel, target_excel)
        print(f"엑셀 파일 복사 완료: {source_excel} -> {target_excel}")
    else:
        print(f"기존 엑셀 파일을 찾을 수 없습니다: {source_excel}")
    
    # 3. 데이터베이스 구조 확인 및 업데이트
    try:
        conn = sqlite3.connect(target_db)
        cursor = conn.cursor()
        
        # 테이블 존재 확인
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [row[0] for row in cursor.fetchall()]
        print(f"현재 테이블 목록: {tables}")
        
        # QA 데이터 개수 확인
        if 'qa_data' in tables:
            cursor.execute("SELECT COUNT(*) FROM qa_data")
            qa_count = cursor.fetchone()[0]
            print(f"QA 데이터 개수: {qa_count}")
        
        # 식단 데이터 개수 확인
        if 'meals' in tables:
            cursor.execute("SELECT COUNT(*) FROM meals")
            meals_count = cursor.fetchone()[0]
            print(f"식단 데이터 개수: {meals_count}")
        
        # 공지사항 데이터 개수 확인
        if 'notices' in tables:
            cursor.execute("SELECT COUNT(*) FROM notices")
            notices_count = cursor.fetchone()[0]
            print(f"공지사항 데이터 개수: {notices_count}")
        
        conn.close()
        
    except Exception as e:
        print(f"데이터베이스 확인 중 오류: {e}")
    
    print("데이터 마이그레이션 완료!")

def update_qa_data_from_excel():
    """엑셀 파일에서 QA 데이터 업데이트"""
    excel_file = "와석초 카카오톡 챗봇 개발을 위한 질문과 답변(0702).xlsx"
    
    if not os.path.exists(excel_file):
        print(f"엑셀 파일을 찾을 수 없습니다: {excel_file}")
        return
    
    try:
        conn = sqlite3.connect("school_data.db")
        cursor = conn.cursor()
        
        total_added = 0
        
        wb = openpyxl.load_workbook(excel_file)
        for sheet_name in wb.sheetnames:
            print(f"{sheet_name} 시트 처리 중...")
            ws = wb[sheet_name]
            
            if sheet_name in ['초등', '유치원']:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 3:
                        question = row[0]  # 질문 예시
                        answer = row[1]    # 답변
                        additional_answer = row[2] if row[2] else ''
                        
                        if question and answer:
                            cursor.execute("SELECT id FROM qa_data WHERE question = ? AND category = ?", (question, sheet_name))
                            existing = cursor.fetchone()
                            if not existing:
                                cursor.execute(
                                    "INSERT INTO qa_data (question, answer, additional_answer, category) VALUES (?, ?, ?, ?)",
                                    (question, answer, additional_answer, sheet_name)
                                )
                                total_added += 1
                                print(f"추가됨: {question[:30]}...")
                            else:
                                print(f"중복됨: {question[:30]}...")
            else:
                # 기타 시트 처리
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 2:
                        question = row[0]  # 첫 번째 컬럼
                        answer = row[1]    # 두 번째 컬럼
                        additional_answer = ''
                        if len(row) > 2:
                            additional_parts = [str(col) for col in row[2:] if col]
                            additional_answer = '\n'.join(additional_parts)
                        
                        if question and answer:
                            cursor.execute("SELECT id FROM qa_data WHERE question = ? AND category = ?", (str(question), sheet_name))
                            existing = cursor.fetchone()
                            if not existing:
                                cursor.execute(
                                    "INSERT INTO qa_data (question, answer, additional_answer, category) VALUES (?, ?, ?, ?)",
                                    (str(question), str(answer), additional_answer, sheet_name)
                                )
                                total_added += 1
                                print(f"추가됨: {str(question)[:30]}...")
                            else:
                                print(f"중복됨: {str(question)[:30]}...")
        
        conn.commit()
        conn.close()
        
        print(f"QA 데이터 업데이트 완료 - 총 {total_added}개 추가됨")
        
    except Exception as e:
        print(f"QA 데이터 업데이트 중 오류: {e}")

if __name__ == "__main__":
    migrate_data()
    update_qa_data_from_excel() 