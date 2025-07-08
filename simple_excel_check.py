import openpyxl

# 엑셀 파일 로드
wb = openpyxl.load_workbook('와석초 카카오톡 챗봇 개발을 위한 질문과 답변(0702).xlsx')

# 결과를 파일로 저장
with open('excel_info.txt', 'w', encoding='utf-8') as f:
    f.write("=== 엑셀 시트 목록 ===\n")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        f.write(f"{i}. {sheet_name}\n")
    
    f.write(f"\n총 {len(wb.sheetnames)}개 시트\n")
    
    # 첫 번째 시트의 구조 확인
    if wb.sheetnames:
        first_sheet = wb[wb.sheetnames[0]]
        f.write(f"\n=== {wb.sheetnames[0]} 시트 구조 ===\n")
        f.write(f"행 수: {first_sheet.max_row}\n")
        f.write(f"열 수: {first_sheet.max_column}\n")
        
        # 첫 3행 출력
        for row in range(1, min(4, first_sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(6, first_sheet.max_column + 1)):
                cell_value = first_sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            f.write(f"행 {row}: {row_data}\n")

wb.close()
print("엑셀 정보가 excel_info.txt 파일에 저장되었습니다.") 