import sqlite3
import openpyxl
import os
from datetime import datetime

def create_database():
    """데이터베이스 및 테이블 생성"""
    conn = sqlite3.connect('school_data.db')
    cursor = conn.cursor()
    
    # 기존 테이블 삭제 (재생성)
    cursor.execute('DROP TABLE IF EXISTS qa_data')
    
    # QA 데이터 테이블 생성
    cursor.execute('''
        CREATE TABLE qa_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            question TEXT NOT NULL,
            answer TEXT NOT NULL,
            link TEXT,
            image_reference TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()
    print("데이터베이스 및 테이블이 생성되었습니다.")

def migrate_excel_data():
    """엑셀 파일의 모든 시트 데이터를 데이터베이스로 마이그레이션"""
    excel_file = "와석초 카카오톡 챗봇 개발을 위한 질문과 답변(0702).xlsx"
    
    if not os.path.exists(excel_file):
        print(f"엑셀 파일을 찾을 수 없습니다: {excel_file}")
        return
    
    conn = sqlite3.connect('school_data.db')
    cursor = conn.cursor()
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        total_inserted = 0
        
        # 1. 초등 시트 처리
        print("=== 초등 시트 처리 중 ===")
        ws_elementary = wb['초등']
        elementary_count = 0
        
        for row_num in range(2, ws_elementary.max_row + 1):  # 헤더 제외
            row_data = []
            for col_num in range(1, 11):  # 10열까지
                cell_value = ws_elementary.cell(row=row_num, column=col_num).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            # 빈 행 건너뛰기
            if not any(row_data[1:3]):  # 질문과 답변이 모두 비어있으면
                continue
                
            question = row_data[1].strip()
            answer = row_data[2].strip()
            
            if question and answer and question != "질문 예시":
                # 링크 추출 (6번째 열)
                link = row_data[5].strip() if len(row_data) > 5 else ""
                
                cursor.execute('''
                    INSERT INTO qa_data (category, question, answer, link)
                    VALUES (?, ?, ?, ?)
                ''', ('초등', question, answer, link))
                elementary_count += 1
        
        print(f"초등 시트: {elementary_count}개 데이터 삽입")
        total_inserted += elementary_count
        
        # 2. 유치원 시트 처리
        print("=== 유치원 시트 처리 중 ===")
        ws_kindergarten = wb['유치원']
        kindergarten_count = 0
        
        for row_num in range(2, ws_kindergarten.max_row + 1):  # 헤더 제외
            row_data = []
            for col_num in range(1, 11):  # 10열까지
                cell_value = ws_kindergarten.cell(row=row_num, column=col_num).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            # 빈 행 건너뛰기
            if not any(row_data[1:3]):  # 질문과 답변이 모두 비어있으면
                continue
                
            question = row_data[1].strip()
            answer = row_data[2].strip()
            
            if question and answer and question != "질문 예시":
                # 링크 추출 (6번째 열)
                link = row_data[5].strip() if len(row_data) > 5 else ""
                
                cursor.execute('''
                    INSERT INTO qa_data (category, question, answer, link)
                    VALUES (?, ?, ?, ?)
                ''', ('유치원', question, answer, link))
                kindergarten_count += 1
        
        print(f"유치원 시트: {kindergarten_count}개 데이터 삽입")
        total_inserted += kindergarten_count
        
        # 3. 첨부 사진 파일 시트 처리
        print("=== 첨부 사진 파일 시트 처리 중 ===")
        ws_images = wb['첨부 사진 파일']
        image_count = 0
        
        # 첨부 사진 파일 시트에서 의미있는 데이터 추출
        image_references = []
        for row_num in range(1, ws_images.max_row + 1):
            for col_num in range(1, ws_images.max_column + 1):
                cell_value = ws_images.cell(row=row_num, column=col_num).value
                if cell_value and str(cell_value).strip():
                    cell_text = str(cell_value).strip()
                    # 의미있는 텍스트만 추출 (빈 문자열이나 공백이 아닌 경우)
                    if len(cell_text) > 2 and not cell_text.isdigit():
                        image_references.append(cell_text)
        
        # 중복 제거 및 정리
        unique_references = list(set(image_references))
        
        # 이미지 참조 정보를 QA 데이터로 추가
        for ref in unique_references:
            if ref and len(ref) > 2:
                cursor.execute('''
                    INSERT INTO qa_data (category, question, answer, image_reference)
                    VALUES (?, ?, ?, ?)
                ''', ('첨부파일', f"{ref} 관련 정보", f"{ref} 이미지 파일 참조", ref))
                image_count += 1
        
        print(f"첨부 사진 파일 시트: {image_count}개 참조 정보 삽입")
        total_inserted += image_count
        
        conn.commit()
        print(f"\n총 {total_inserted}개의 데이터가 성공적으로 마이그레이션되었습니다.")
        
        # 마이그레이션 결과 확인
        cursor.execute('SELECT category, COUNT(*) FROM qa_data GROUP BY category')
        results = cursor.fetchall()
        print("\n=== 마이그레이션 결과 ===")
        for category, count in results:
            print(f"{category}: {count}개")
        
        wb.close()
        
    except Exception as e:
        print(f"마이그레이션 중 오류 발생: {e}")
        conn.rollback()
    finally:
        conn.close()

def verify_migration():
    """마이그레이션 결과 확인"""
    conn = sqlite3.connect('school_data.db')
    cursor = conn.cursor()
    
    print("\n=== 데이터베이스 내용 확인 ===")
    
    # 전체 데이터 수 확인
    cursor.execute('SELECT COUNT(*) FROM qa_data')
    total_count = cursor.fetchone()[0]
    print(f"전체 데이터 수: {total_count}개")
    
    # 카테고리별 데이터 수 확인
    cursor.execute('SELECT category, COUNT(*) FROM qa_data GROUP BY category')
    category_counts = cursor.fetchall()
    for category, count in category_counts:
        print(f"{category}: {count}개")
    
    # 샘플 데이터 확인
    print("\n=== 샘플 데이터 (각 카테고리별 3개씩) ===")
    for category, _ in category_counts:
        cursor.execute('SELECT question, answer FROM qa_data WHERE category = ? LIMIT 3', (category,))
        samples = cursor.fetchall()
        print(f"\n[{category}]")
        for i, (question, answer) in enumerate(samples, 1):
            print(f"{i}. Q: {question[:50]}...")
            print(f"   A: {answer[:50]}...")
    
    conn.close()

if __name__ == "__main__":
    print("와석초 카카오톡 챗봇 데이터 마이그레이션 시작")
    print("=" * 50)
    
    # 1. 데이터베이스 생성
    create_database()
    
    # 2. 엑셀 데이터 마이그레이션
    migrate_excel_data()
    
    # 3. 결과 확인
    verify_migration()
    
    print("\n마이그레이션이 완료되었습니다!") 